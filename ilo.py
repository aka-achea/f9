

import subprocess

import configparser
import paramiko
from pprint import pprint
import pandas as pd
import csv
import openpyxl 
import webbrowser

confile = r'E:\ptool\ilo.ini'

config = configparser.ConfigParser()
config.read(confile)
# host = config['login']['host']
adm = config['login']['adm']
# pword = config['login']['pword']
newuser = config['user']['user']
newuserpd = config['user']['password']
logfile = config['setting']['log']
xls = config['setting']['xlsx']


paramiko.util.log_to_file(logfile,level='DEBUG')


def create_ilo_user(host,adm,pword):
    '''Create iLO user, get SN, Model'''
    sn,model = '',''
    try:
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.client.AutoAddPolicy)
        print(host,pword)
        client.connect(hostname=host,username=adm,password=pword,
                        allow_agent=False,look_for_keys=False)
        cmd = f'create /map1/accounts1 username={newuser} password={newuserpd} group=admin,config,oemHPE_rc,oemHPE_power,oemHPE_vm'
        client.exec_command(cmd)

        client.connect(hostname=host,username=newuser,password=newuserpd,
                        allow_agent=False,look_for_keys=False)
        stdin, stdout, stderr = client.exec_command('show system1 name')
        output = stdout.read().decode('utf-8').split('\r\n')
        for x in output:
            if 'name=' in x:
                model = x.split('name=')[1]
        stdin, stdout, stderr = client.exec_command('show system1 number')
        output = stdout.read().decode('utf-8').split('\r\n')
        for x in output:
            if 'number=' in x:
                sn = x.split('number=')[1]  
   
    finally:
        client.close()
    return sn,model


def setup_ilo():
    wb = openpyxl.load_workbook(xls)
    sheet = wb['hw']
    for x in range(2,sheet.max_row+1):
        if sheet.cell(row=x,column=9).value != 'Y':
            host = sheet.cell(row=x,column=10).value
            pword = sheet.cell(row=x,column=11).value
            sn,model = create_ilo_user(host,adm,pword)
            # print(sn,model)
            sheet.cell(row=x,column=8).value = sn
            sheet.cell(row=x,column=6).value = model
            url = f'https://{host}'
            webbrowser.open(url)
    wb.save(xls)


if __name__ == "__main__":
    setup_ilo()